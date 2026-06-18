#!/usr/bin/env python3
"""Deterministic earnings-PROJECTION + CHART generator for JSE research.

Companion to tools/build_model.py. Where build_model.py emits the full live-formula
scenario workbook, this tool emits a chart-led projection book that answers one
question visually: how do earnings track in a BULL, NEUTRAL or BEAR market?

It reads the SAME version-controlled inputs (no new source of truth):

  assumptions/<slug>.json   per-company FY anchor (sacred actuals) + scenario drivers
  macro/latest.json         shared dated macro (read only for the Cover context line)

For each company it computes, per scenario and per forward year, four projected
series and ranks the scenarios into Bull / Neutral / Bear by terminal-year headline
earnings (so it works for any scenario keys, not just A/Base/B):

  - Revenue (R'm)
  - Trading / operating profit (R'm)
  - Headline earnings (R'm)
  - Dividend per share (cents)

The financial logic mirrors build_model.py exactly (retail-owned and retail-wholesale
templates) so the two tools never disagree. Forward figures are analyst estimates and
are emitted as numbers (cached) so the native Excel charts render on open; all are
marked (e). FY-anchor figures derive from the reported actuals in the anchor block.

ALL currency figures are in millions of rands (R'm). Per-share figures are in cents.

Usage:
  python3 tools/build_projection.py shoprite --out projection.xlsx
  python3 tools/build_projection.py shoprite spar --title "SHP & SPP - Earnings Projections" \\
      --out compare_projection.xlsx
"""
import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import report_style as rs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Bull / Neutral / Bear stance colours (house palette).
STANCE_FILL = {"Bull": rs.GREEN, "Neutral": rs.GREY, "Bear": rs.RED}
STANCE_ORDER = ["Bull", "Neutral", "Bear"]

# The four projected series. (key, label, number_format, is_currency_rm)
SERIES_SPECS = [
    ("revenue", "Revenue (R'm)", rs.FMT_NUM, True),
    ("op_profit", "Trading / operating profit (R'm)", rs.FMT_NUM, True),
    ("hl_earnings", "Headline earnings (R'm)", rs.FMT_NUM, True),
    ("dps", "Dividend per share (cents)", rs.FMT_CENT, False),
]


def load_json(path):
    with open(path) as f:
        return json.load(f)


# --------------------------------------------------------------- projection maths
def project_company(comp):
    """Return (periods, scenarios) where:
       periods = [anchor_label, year1, year2, ...]
       scenarios = list of dicts: {key,label,stance,series:{metric:[vals per period]}}
    The first value of every series is the FY-anchor figure (shared across scenarios).
    """
    a = comp["anchor"]
    template = comp["template"]
    # forward year labels (assume identical across scenarios, as in build_model)
    year_labels = list(comp["scenarios"][0]["years"].keys())
    periods = [comp["anchor_label"]] + year_labels

    # --- FY anchor figures (derived from reported actuals) -------------------
    if template == "retail-owned":
        anc_sale = a["sale_of_merchandise_rm"]
        anc_rev = anc_sale * a["group_rev_ratio"]
        anc_op = anc_sale * a["trading_margin"]
        anc_earn = a["heps_actual_c"] * a["eff_diluted_shares_m"] / 100.0
        anc_dps = a["heps_actual_c"] / a["div_cover"]
    elif template == "retail-wholesale":
        anc_to = a["turnover_rm"]
        anc_rev = anc_to
        anc_op = anc_to * a["op_margin"]
        anc_earn = a["heps_actual_c"] * a["shares_m"] / 100.0
        anc_dps = a.get("dps_c", 0)
    else:
        raise SystemExit(f"Unknown template '{template}' for {comp['slug']}")
    anchor_vals = {"revenue": anc_rev, "op_profit": anc_op,
                   "hl_earnings": anc_earn, "dps": anc_dps}

    scenarios = []
    for sc in comp["scenarios"]:
        series = {k: [anchor_vals[k]] for k, *_ in SERIES_SPECS}
        if template == "retail-owned":
            prior_sale = a["sale_of_merchandise_rm"]
            for yr in year_labels:
                d = sc["years"][yr]
                sale = prior_sale * (1 + d["infl"] + d["vol"])
                rev = sale * a["group_rev_ratio"]
                op = sale * d["trading_margin"]
                pbt = op - d["net_finance_costs_rm"]
                earn = pbt * (1 - a["tax_rate"])
                heps_c = earn / a["eff_diluted_shares_m"] * 100.0
                dps = heps_c / a["div_cover"]
                series["revenue"].append(rev)
                series["op_profit"].append(op)
                series["hl_earnings"].append(earn)
                series["dps"].append(dps)
                prior_sale = sale
        else:  # retail-wholesale
            prior_to = a["turnover_rm"]
            for yr in year_labels:
                d = sc["years"][yr]
                to = prior_to * (1 + d["turnover_growth"])
                op = to * d["op_margin"]
                earn = d["heps_c"] * a["shares_m"] / 100.0
                series["revenue"].append(to)
                series["op_profit"].append(op)
                series["hl_earnings"].append(earn)
                series["dps"].append(d.get("dps_c", 0))
                prior_to = to
        scenarios.append({"key": sc["key"], "label": sc["label"], "series": series})

    # --- classify scenarios into Bull / Neutral / Bear ----------------------
    # rank by terminal-year headline earnings (highest = Bull, lowest = Bear)
    ranked = sorted(scenarios, key=lambda s: s["series"]["hl_earnings"][-1], reverse=True)
    n = len(ranked)
    for i, s in enumerate(ranked):
        if i == 0:
            s["stance"] = "Bull"
        elif i == n - 1 and n >= 2:
            s["stance"] = "Bear"
        else:
            s["stance"] = "Neutral"
    # keep original input order for display stability
    return periods, scenarios, anchor_vals


# ------------------------------------------------------------------- sheet: Cover
def build_cover(wb, companies, macro, title, subtitle):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    rs.xl_title_block(ws, title, subtitle, span=8)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 16
    r = 4
    intro = [
        ("HOW TO READ THIS BOOK", "h"),
        ("Each company has a DATA tab (the projected numbers) and a CHARTS tab "
         "(the same numbers as graphs). Scenarios are ranked into BULL (best), "
         "NEUTRAL and BEAR (worst) by projected headline earnings in the final "
         "forward year.", "p"),
        ("All currency figures are in millions of rands (R'm). Per-share figures "
         "are in cents. Forward figures are analyst estimates (e), not company "
         "guidance; the FY anchor derives from reported actuals.", "p"),
        ("Generated by tools/build_projection.py from assumptions/<slug>.json "
         "(+ macro/latest.json for context). Do not hand-edit; re-run the generator.", "p"),
        ("", ""),
        ("SCENARIOS (ranked)", "h"),
    ]
    for text, kind in intro:
        c = ws.cell(r, 2, text)
        if kind == "h":
            rs.xl_cell(c, bold=True, size=11, color=rs.NAVY)
        elif kind == "p":
            rs.xl_cell(c, size=9, wrap=True)
            ws.row_dimensions[r].height = 40
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1
    # per-company scenario legend
    for comp in companies:
        rs.xl_cell(ws.cell(r, 2, f"{comp['name']} ({comp['jse_code']})  -  anchor {comp['anchor_label']}"),
                   bold=True, size=10, color=rs.WHITE, fill=rs.NAVY)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1
        for stance in STANCE_ORDER:
            sc = next((s for s in comp["_scenarios"] if s["stance"] == stance), None)
            if not sc:
                continue
            rs.xl_cell(ws.cell(r, 2, stance), bold=True, size=9, color=rs.WHITE,
                       fill=STANCE_FILL[stance], align="center", border=True)
            rs.xl_cell(ws.cell(r, 3, sc["label"]), size=9, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
            r += 1
        if comp.get("rationale"):
            rs.xl_cell(ws.cell(r, 2, comp["rationale"]), size=8, italic=True, wrap=True, fill=rs.YELLOW)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            ws.row_dimensions[r].height = 80
            r += 1
        r += 1
    if macro:
        rs.xl_cell(ws.cell(r, 2, f"Macro context: {macro.get('title','')} (as at {macro.get('as_at','')})"),
                   size=8, italic=True, wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 2
    rs.xl_cell(ws.cell(r, 2, "DISCLAIMER"), bold=True, size=11, color=rs.NAVY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8); r += 1
    rs.xl_cell(ws.cell(r, 2, rs.SACRED_NOTE + " " + rs.DISCLAIMER), size=8, italic=True, wrap=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 70
    r += 2
    src = "; ".join(f"{c['jse_code']}: {c['anchor'].get('source','assumptions/'+c['slug']+'.json')}"
                    for c in companies)
    rs.xl_cell(ws.cell(r, 2, "Sources: " + src), size=8, italic=True, wrap=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 50


# --------------------------------------------------------------- sheet: Data + Charts
def build_company_sheets(wb, comp):
    periods = comp["_periods"]
    scenarios = comp["_scenarios"]
    # order scenarios Bull, Neutral, Bear for chart legibility
    ordered = sorted(scenarios, key=lambda s: STANCE_ORDER.index(s["stance"]))
    code = comp["jse_code"]
    ws = wb.create_sheet(f"{code} Data")
    ws.sheet_view.showGridLines = False
    rs.xl_title_block(
        ws, f"{comp['name']} ({code}) - EARNINGS PROJECTIONS",
        "R'm unless stated (per-share figures in cents). Forward figures are analyst "
        "estimates (e); FY anchor from reported actuals.", span=2 + len(periods))
    ncol = len(periods)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    for i in range(ncol):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14

    # remember the cell ranges of each metric block for charting
    blocks = {}  # metric_key -> {"hdr_row":, "first":, "last":, "cat_first":3,"cat_last":}
    r = 4
    for mkey, mlabel, mfmt, is_rm in SERIES_SPECS:
        # section header row: metric name + period labels
        rs.xl_cell(ws.cell(r, 2, mlabel), bold=True, size=10, color=rs.WHITE, fill=rs.NAVY, border=True)
        for i, p in enumerate(periods):
            rs.xl_cell(ws.cell(r, 3 + i, p), bold=True, size=9, color=rs.WHITE,
                       fill=rs.MIDGREY, align="center", border=True)
        hdr_row = r
        r += 1
        first = r
        for sc in ordered:
            rs.xl_cell(ws.cell(r, 2, sc["stance"]), bold=True, size=9, color=rs.WHITE,
                       fill=STANCE_FILL[sc["stance"]], border=True)
            for i, v in enumerate(sc["series"][mkey]):
                rs.xl_cell(ws.cell(r, 3 + i, round(v, 1)), size=9, fmt=mfmt,
                           align="center", border=True)
            r += 1
        last = r - 1
        blocks[mkey] = {"hdr": hdr_row, "first": first, "last": last,
                        "cat_first": 3, "cat_last": 3 + ncol - 1}
        r += 1  # gap between metric blocks

    # ---- charts sheet ----
    cs = wb.create_sheet(f"{code} Charts")
    cs.sheet_view.showGridLines = False
    rs.xl_title_block(
        cs, f"{comp['name']} ({code}) - PROJECTION CHARTS",
        "Bull / Neutral / Bear across the FY anchor and forward years. R'm unless "
        "stated; per-share in cents. Forward figures (e).", span=12)

    anchor_pos = [("B", 4), ("J", 4), ("B", 22), ("J", 22)]
    for idx, (mkey, mlabel, mfmt, is_rm) in enumerate(SERIES_SPECS):
        blk = blocks[mkey]
        # headline earnings as a LINE chart (trajectory); the rest clustered columns
        if mkey == "hl_earnings":
            chart = LineChart()
            chart.style = 2
        else:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.style = 10
        chart.title = mlabel
        chart.height = 8.2
        chart.width = 15.5
        chart.y_axis.title = "R'm" if is_rm else "cents"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        cats = Reference(ws, min_col=blk["cat_first"], max_col=blk["cat_last"], min_row=blk["hdr"])
        # one series per stance row, coloured by stance
        for row in range(blk["first"], blk["last"] + 1):
            stance = ws.cell(row, 2).value
            ser = Series(Reference(ws, min_col=blk["cat_first"], max_col=blk["cat_last"], min_row=row),
                         title=stance)
            colour = STANCE_FILL.get(stance, rs.GREY)
            if mkey == "hl_earnings":
                ser.graphicalProperties.line.solidFill = colour
                ser.graphicalProperties.line.width = 28000
                ser.smooth = False
            else:
                ser.graphicalProperties.solidFill = colour
                ser.graphicalProperties.line.solidFill = colour
            chart.series.append(ser)
        chart.set_categories(cats)
        col_letter, row_anchor = anchor_pos[idx]
        cs.add_chart(chart, f"{col_letter}{row_anchor}")
    return ws, cs


# ----------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Build the charted earnings-projection workbook.")
    ap.add_argument("slugs", nargs="+", help="company slug(s) with an assumptions/<slug>.json")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    ap.add_argument("--title", default=None)
    ap.add_argument("--subtitle", default=None)
    args = ap.parse_args()

    macro_path = os.path.join(ROOT, "macro", "latest.json")
    macro = load_json(macro_path) if os.path.exists(macro_path) else None

    companies = []
    for slug in args.slugs:
        path = os.path.join(ROOT, "assumptions", f"{slug}.json")
        if not os.path.exists(path):
            raise SystemExit(f"Missing {path}. Onboard the company / create its assumptions first.")
        comp = load_json(path)
        periods, scenarios, anchor = project_company(comp)
        comp["_periods"] = periods
        comp["_scenarios"] = scenarios
        companies.append(comp)

    title = args.title or (
        f"{companies[0]['name']} - Earnings Projections" if len(companies) == 1
        else " & ".join(c["jse_code"] for c in companies) + " - Earnings Projections")
    subtitle = args.subtitle or "Bull / Neutral / Bear scenarios  -  all currency in R'm"

    wb = Workbook()
    build_cover(wb, companies, macro, title, subtitle)
    for comp in companies:
        build_company_sheets(wb, comp)

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}")
    for comp in companies:
        line = []
        for sc in sorted(comp["_scenarios"], key=lambda s: STANCE_ORDER.index(s["stance"])):
            term = sc["series"]["hl_earnings"][-1]
            line.append(f"{sc['stance']}={term:,.0f} R'm")
        print(f"  {comp['jse_code']} terminal-year headline earnings: " + "; ".join(line))


if __name__ == "__main__":
    main()
